import global_variables
import linkedin_scrapers
import web_common
import os

from openpyxl import load_workbook

def check_if_company_exists_in_database():
    try:
        companies_database_path = global_variables.obsidian_companies_database_path
        current_url = web_common.extract_home_link(global_variables.current_url)

        return value_exists_in_file(
            companies_database_path,
            current_url,
            'Companies (ALL)',
            '✅ IS IN COMPANIES DATABASE\n',
            '❌ NOT IN COMPANIES DATABASE\n'
        )

    except Exception as e:
        print('❌ Error: {e}')

def check_if_linkedin_company_exists_in_database():
    try:
        companies_database_path = global_variables.obsidian_companies_database_path

        if web_common.company_url_specific_comparasion(global_variables.current_url):
            current_url = web_common.normalize_company_url(global_variables.current_url)
        else:
            current_url = global_variables.current_url

        return value_exists_in_file(
            companies_database_path,
            current_url,
            'Companies (ALL)',
            '✅ IS IN COMPANIES DATABASE\n',
            '❌ NOT IN COMPANIES DATABASE\n'
        )

    except Exception as e:
        print('❌ Error: {e}')

def check_if_careers_page_exists_in_database():
    try:
        companies_database_path = global_variables.obsidian_companies_database_path
        careers_url = global_variables.current_url

        return value_exists_in_file(
            companies_database_path,
            careers_url, 'Careers Pages',
            '✅ IS IN CAREERS PAGE DATABASE\n',
            '❌ NOT IN CAREERS PAGE DATABASE\n'
        )

    except Exception as e:
        print(f'\n❌ Error: {e}')

def append_row(target_file, values, target_sheet, print_if_success=True):
    try:
        workbook = load_workbook(target_file)
        sheet = workbook[target_sheet]

        # Append row with values
        sheet.append(values)

        # Save changes
        workbook.save(target_file)

        if print_if_success:
            print('\n✅ File appended successfully!')

    except Exception as e:
        print(f'\n❌ Error: {e}')

def value_exists_in_file(file_path, value, target_sheet, message_found='', message_not_found=''):
    try:
        # Load the workbook
        workbook = load_workbook(file_path)

        # Iterate through all sheets and cells to find the value
        for sheet in workbook.sheetnames:
            ws = workbook[target_sheet]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == value:
                        if message_found:
                            print(message_found)
                        return True

        # Value not found
        if message_not_found:
            print(message_not_found)
        return False

    except Exception as e:
        print(f"\n❌ Error while checking if the value exists in the database: {e}")
        return False

def append_company_to_database():
    try:
        companies_database_path = global_variables.obsidian_companies_database_path

        if not global_variables.current_url.startswith('https://www.linkedin.com/company/'):
            print('\n❌ NOT A VALID LINKEDIN COMPANY URL')
            return False

        if web_common.company_url_specific_comparasion(global_variables.current_url):
            current_url = web_common.normalize_company_url(global_variables.current_url)
        else:
            current_url = global_variables.current_url

        if value_exists_in_file(companies_database_path, current_url, "Companies (ALL)"):
            print('\n✅ COMPANY ALREADY IN THE DATABASE')
            return False

        global_variables.active_input = True
        os.system('cls||clear')

        company_name = linkedin_scrapers.scrape_linkedin_company_name()
        about = linkedin_scrapers.scrape_linkedin_company_description()
        location = linkedin_scrapers.scrape_linkedin_company_location()
        industry = linkedin_scrapers.scrape_linkedin_company_industry()

        while True:
            careers_url = input('➡️ Enter the Careers Page URL (blank to enter the Company URL): ')

            if web_common.is_valid_url(careers_url):
                if append_careers_page_to_database(careers_url):
                    company_url = web_common.extract_home_link(careers_url)
                    break

            elif not careers_url:
                while True:

                    company_url = input('➡️ Enter the Company URL: ')
                    if web_common.is_valid_url(company_url):
                        company_url = web_common.extract_home_link(company_url)
                        break
                    else:
                        print('\n❌ The URL that you entered is not valid. Please try again.')

                break

            else:
                print('\n❌ The URL that you entered is not valid. Please try again.')

        tags = input('➡️ Enter company tags: ')

        values = [company_name, current_url, about, industry, location, company_url, tags]
        append_row(companies_database_path, values, 'Companies (ALL)')

    except Exception as e:
        print(f'❌ Error while appending company to database: {e}')

    global_variables.active_input = False

def append_careers_page_to_database(careers_url=None):
    try:
        global_variables.active_input = True
        if not careers_url:
            os.system('cls||clear')

        companies_database_path = global_variables.obsidian_companies_database_path

        if not careers_url:
            careers_url = global_variables.current_url
            home_url = web_common.extract_home_link(careers_url)
            url = web_common.extract_home_link(input(f'➡️ Enter URL (default - {home_url}): ')) or home_url
            print_if_success = True
        else:
            url = web_common.extract_home_link(careers_url)
            print_if_success = False

        if value_exists_in_file(companies_database_path, careers_url, "Careers Pages"):
            print('\n✅ URL ALREADY IN CAREERS PAGES')
            global_variables.active_input = False
            return False

        department = input('➡️ Enter careers department (default - General): ') or 'General'
        office = input('➡️ Enter careers office (default - All): ') or 'All'
        xpath = input('➡️ Enter careers Next Button XPath: ')

        values = [url, careers_url, department, office, xpath]
        append_row(companies_database_path, values, 'Careers Pages', print_if_success=print_if_success)

        if print_if_success:
            global_variables.active_input = False

        return True

    except Exception as e:
        print(f'\n❌ Error while appending careers page to the database: {e}')
        global_variables.active_input = False
        return False

def get_rows(file_path, target_sheet):
    try:
        # Load the workbook
        wb = load_workbook(file_path)

        # Check if "Careers Page" sheet exists
        if target_sheet in wb.sheetnames:
            # Get the "Careers Page" sheet
            sheet = wb[target_sheet]
            return sheet.iter_rows(values_only=True)

        else:
            print(f"\nError: Sheet {target_sheet} not found in the Excel file.")

    except Exception as e:
        print(f"\nError: {e}")

    return []
